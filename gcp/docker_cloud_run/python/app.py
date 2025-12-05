import solara
import pandas as pd
from io import BytesIO
import ipywidgets as widgets
from IPython.display import display
from openpyxl import Workbook
from openpyxl.styles import PatternFill

class DynamicTable:
    def __init__(self):
        self.rows = []
        self.default = widgets.IntText(description="Default:", value=0, layout=widgets.Layout(width="200px"))
        self.default.observe(lambda c: [cell.set_trait('value', str(c['new'])) for r in self.rows for cell in r[0] if not cell.value.strip()], 'value')
        self.container = widgets.VBox()

        add_btn = widgets.Button(description="Add Row", button_style="success")
        add_btn.on_click(lambda b: self.add_row())

        # Inject CSS for red, green, yellow cells
        display(widgets.HTML("""
        <style>
        .red-cell input {
            background-color: red !important;
        }
        .green-cell input {
            background-color: lightgreen !important;
        }
        .yellow-cell input {
            background-color: yellow !important;
        }
        </style>
        """))

        display(self.default, add_btn, self.container,
                solara.FileDownload(self.generate_excel, filename="table.xlsx", label="Download Excel"))

    def add_row(self):
        cells = [widgets.Text(description=f"Col {i+1}", value=str(self.default.value)) for i in range(5)]

        row_index = len(self.rows)

        # First row: red in col 2, green in col 3
        if row_index == 0:
            cells[1].add_class('red-cell')
            cells[2].add_class('green-cell')

        # Second row: yellow directly below the red cell
        if row_index == 1:
            cells[1].add_class('yellow-cell')

        for c in cells:
            c.observe(lambda ch, w=c: w.set_trait('value', str(self.default.value)) if not ch['new'].strip() else None, 'value')

        rm = widgets.Button(description="Remove", button_style="danger", layout=widgets.Layout(width="80px"))
        hbox = widgets.HBox(cells + [rm])
        rm.on_click(lambda b: (self.rows.remove(next(r for r in self.rows if r[1] == b)),
                               setattr(self.container, 'children', [r[2] for r in self.rows])))

        self.rows.append((cells, rm, hbox))
        self.container.children = [r[2] for r in self.rows]

    def generate_excel(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "Data"

        color_map = {
            'red-cell': 'FF0000',
            'green-cell': '90EE90',
            'yellow-cell': 'FFFF00'
        }

        for r_idx, row in enumerate(self.rows, start=1):
            for c_idx, cell in enumerate(row[0], start=1):
                ws.cell(row=r_idx, column=c_idx, value=cell.value)
                # Detect CSS classes from frontend widgets
                if hasattr(cell, '_dom_classes'):
                    for cls in cell._dom_classes:
                        if cls in color_map:
                            ws.cell(row=r_idx, column=c_idx).fill = PatternFill(start_color=color_map[cls], end_color=color_map[cls], fill_type="solid")

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.read()

DynamicTable()
